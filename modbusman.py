# ver.11 pridani kontroly ceny dle OTE
# ver.12 zruseni regulace dle HDO; pridani do podminky regulace dle OTE kontrolu v 15. 30. a 45. minute; zavedeni ceny do DB
# ver.13 prenastaveni podprogramu OTE tak, aby vyhodnocoval podminku ve stejnem cyklu jako hlavni program
# ver.14 prenastaveni podprogramu OTE tak, aby drzel rele sepnute po celou dobu platnosti nastavene ceny OTE
# ver.15 zmena volani funkce spinani rele; kod prevzat od vyrobce rele

from pymodbus.client import ModbusSerialClient
import struct
import time
import subprocess
import usbrelay_py

######################################### otevreni a nastaveni serioveho portu
client = ModbusSerialClient(
    port='/dev/ttyUSB0',
    baudrate=9600,
    stopbits=1,
    bytesize=8,
    parity='N',
    timeout=1
)
#########################################

######################################### zadani vychozich hodnot
frequency = 1
UL1 = 1
UL2 = 1
UL3 = 1
IL1 = 1
IL2 = 1
IL3 = 1
PL1 = 1
PL2 = 1
PL3 = 1
impPower = 1
expPower = 1
continueWhile = "yes"
error = ""
errorFrequency = ""
errorUL1 = ""
errorUL2 = ""
errorUL3 = ""
errorIL1 = ""
errorIL2 = ""
errorIL3 = ""
errorPL1 = ""
errorPL2 = ""
errorPL3 = ""
sendErrorMessageEmoog = 1
sendGridRecovery = 0
usbrelay0_1ncon = 0 # vychozi stav rele; usbrelay0_1ncon = 1 zapnuto; z duvodu omezeni povelu pro rele
priceOTE = 0
statusRelay = False
seznamPrice = 0
relay = 1

######################################### uvodni vydefinovani seznamu cen, aby se program spustil
# kdykoliv mimo nacteni noveho souboru v 0:01 viz. podprogram vycteni cen OTE a vypnuti pretoku
# pri stanovene cene
from openpyxl import load_workbook 
from datetime import datetime
todayDate = datetime.today()
todayTime = datetime.today()
hour = datetime.now().hour
minute = datetime.now().minute
day = todayDate.strftime("%d")
month = todayDate.strftime("%m")
year = todayDate.strftime("%y")
wb = load_workbook(f"/home/tomas/OTE/OTEfiles/DT_{day}_{month}_20{year}_CZ.xlsx")
sheet = wb.active
cellsPrice = ['B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B30', 'B31', 'B32', 'B33', 'B34', 'B35',
            'B36', 'B37', 'B38', 'B39', 'B40', 'B41', 'B42', 'B43', 'B44', 'B45', 'B46', 'B47']
seznamPrice = [sheet[cell].value for cell in cellsPrice]
#########################################

######################################### vytvoreni databaze
import sqlite3
conn = sqlite3.connect('/home/tomas/OTE/test11.db')
c = conn.cursor()
c.execute('''
CREATE TABLE IF NOT EXISTS cisla (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    time TEXT,
    frequency FLOAT,
    UL1 FLOAT,
    UL2 FLOAT,
    UL3 FLOAT,
    IL1 FLOAT,
    IL2 FLOAT,
    IL3 FLOAT,
    PL1 FLOAT,
    PL2 FLOAT,
    PL3 FLOAT,
    impPower FLOAT,
    expPower FLOAT,
    priceOTE,
    statusRelay,
    error TEXT
    
    )
''')
conn.commit()
conn.close()
######################################### zacatek programu

while (continueWhile == "yes"):
#for i in range(1, 11):

######################################### nacteni hodnot ze serioveho portu
    if client.connect():
        
        try: # Frequency
            resultFrequency = client.read_input_registers(address=70, count=2, slave=1)
            if resultFrequency.isError():
                resultFrequency = 100
                errorFrequency = "NEZOBRAZUJE SE"
            else:
                resultFrequency = client.read_input_registers(address=70, count=2, slave=1)
                if not resultFrequency.isError():
                    regsFrequency = resultFrequency.registers
#           		print(f"Surova data: {regsFrequency}")
                    bytes_ = struct.pack('>HH', regsFrequency[0], regsFrequency[1])
                    frequency = struct.unpack('>f', bytes_)[0]
#           		print(f"Float hodnota: {frequency}")
                else:
                    print("Chyba pri cteni:", resultFrequency)
                client.close()
                ############################### vypadek a obnoveni site
                sendErrorMessageEmoog = 1 # 1 = po obnoveni site se nastavi aby pri dalsim vypadku odesel dalsi mail
                if sendGridRecovery == 1: # 1 = po vypadku site se nastavi do 1
                    subprocess.run(["python3", "gr.py"])
                    sendGridRecovery = 0 # 0 = zajisti aby se mail poslal po obnove site jen jednou
                ###############################
                if frequency > 51:
                    errorFrequency = "OVERFREQUENCY > 50Hz"
                if frequency < 48:
                    errorFrequency = "UNDERFREQUENCY < 48Hz"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultFrequency = 0
            frequency = 100
            errorFrequency = "Invalid data frequency"
            #print("Frequency:", resultFrequency)
            print("Frequency:", frequency)
#########################################            
        try: # VoltagePhase1toN
            resultVoltagePhase1toN = client.read_input_registers(address=0, count=2, slave=1)
            if resultVoltagePhase1toN.isError():
                resultVoltagePhase1toN = 201
                errorUL1 = "NEZOBRAZUJE SE"
            else:
                resultVoltagePhase1toN = client.read_input_registers(address=0, count=2, slave=1)
                if not resultVoltagePhase1toN.isError():
                    regsVoltagePhase1toN = resultVoltagePhase1toN.registers
#           		print(f"Surova data: {regsVoltagePhase1toN}")
                    bytes_ = struct.pack('>HH', regsVoltagePhase1toN[0], regsVoltagePhase1toN[1])
                    UL1 = struct.unpack('>f', bytes_)[0]
#           		print(f"Float hodnota: {UN1}")
                else:
                    print("Chyba pri cteni:", resultVoltagePhase1toN)
                client.close()
                if UL1 > 255:
                    errorUL1 = "OVERVOLTAGE L1 > 255V"
                if UL1 < 220:
                    errorUL1 = "UNDERVOLTAGE L1 < 220V"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultVoltagePhase1toN = 0
            UL1 = 201
            errorUL1 = "Invalid data UL1"
            #print("resultVoltagePhase1toN:", resultVoltagePhase1toN)
            print("resultVoltagePhase1toN:", UL1)
#########################################            
        try: # VoltagePhase2toN
            resultVoltagePhase2toN = client.read_input_registers(address=1, count=2, slave=1)
            if resultVoltagePhase2toN.isError():
                resultVoltagePhase2toN = 202
                errorUL2 = "NEZOBRAZUJE SE"
            else:
                resultVoltagePhase2toN = client.read_input_registers(address=1, count=2, slave=1)
                if not resultVoltagePhase2toN.isError():
                    regsVoltagePhase2toN = resultVoltagePhase2toN.registers
#           		print(f"Surova data: {regsVoltagePhase1toN}")
                    bytes_ = struct.pack('>HH', regsVoltagePhase2toN[1], regsVoltagePhase2toN[0])
                    UL2 = struct.unpack('>f', bytes_)[0]
#           		print(f"Float hodnota: {UN1}")
                else:
                    print("Chyba pri cteni:", resultVoltagePhase2toN)
                client.close()
                if UL2 > 255:
                    errorUL2 = "OVERVOLTAGE L2 > 255V"
                if UL2 < 220:
                    errorUL2 = "UNDERVOLTAGE L2 < 220V"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultVoltagePhase2toN = 0
            UL2 = 202
            errorUL2 = "Invalid data UL2"
            #print("resultVoltagePhase2toN:", resultVoltagePhase2toN)
            print("resultVoltagePhase2toN:", UL2)
#########################################            
        try: # VoltagePhase3toN
            resultVoltagePhase3toN = client.read_input_registers(address=2, count=2, slave=1)
            if resultVoltagePhase3toN.isError():
                resultVoltagePhase3toN = 203
                errorUL3 = "NEZOBRAZUJE SE"
            else:
                resultVoltagePhase3toN = client.read_input_registers(address=2, count=2, slave=1)
                if not resultVoltagePhase3toN.isError():
                    regsVoltagePhase3toN = resultVoltagePhase3toN.registers
#           		print(f"Surova data: {regsVoltagePhase1toN}")
                    bytes_ = struct.pack('>HH', regsVoltagePhase3toN[0], regsVoltagePhase3toN[1])
                    UL3 = struct.unpack('>f', bytes_)[0]
#           		print(f"Float hodnota: {UN1}")
                else:
                    print("Chyba pri cteni:", resultVoltagePhase3toN)
                client.close()
                if UL3 > 255:
                    errorUL3 = "OVERVOLTAGE L3 > 255V"
                if UL3 < 220:
                    errorUL3 = "UNDERVOLTAGE L3 < 220V"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultVoltagePhase3toN = 0
            UL3 = 203
            errorUL3 = "Invalid data UL3"
            #print("resultVoltagePhase3toN:", resultVoltagePhase3toN)
            print("resultVoltagePhase3toN:", UL3)
#########################################
        try: # CurrentPhase1
            resultCurrentPhase1 = client.read_input_registers(address=6, count=2, slave=1)
            if resultCurrentPhase1.isError():
                resultCurrentPhase1 = 301
#               errorUL1 = "NEZOBRAZUJE SE"
            else:
                resultCurrentPhase1 = client.read_input_registers(address=6, count=2, slave=1)
                if not resultCurrentPhase1.isError():
                    regsCurrentPhase1 = resultCurrentPhase1.registers
#                   print(f"Surova data IL1: {regsCurrentPhase1}")
                    bytes_ = struct.pack('>HH', regsCurrentPhase1[0], regsCurrentPhase1[1])
                    IL1 = struct.unpack('>f', bytes_)[0]
#                   print(f"Float hodnota IL1: {IL1}")
                else:
                    print("Chyba pri cteni:", resultCurrentPhase1)
                client.close()
#               if UL1 > 255:
#                   errorUL1 = "OVERVOLTAGE L1 > 255V"
#               if UL1 < 220:
#                   errorUL1 = "UNDERVOLTAGE L1 < 220V"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultCurrentPhase1 = 0
            IL1 = 301
            errorIL1 = "Invalid data IL1"
            #print("resultCurrentPhase1:", resultCurrentPhase1)
            print("resultCurrentPhase1:", IL1)            
#########################################
        try: # CurrentPhase2
            resultCurrentPhase2 = client.read_input_registers(address=8, count=2, slave=1)
            if resultCurrentPhase2.isError():
                resultCurrentPhase2 = 302
#               errorUL1 = "NEZOBRAZUJE SE"
            else:
                resultCurrentPhase2 = client.read_input_registers(address=8, count=2, slave=1)
                if not resultCurrentPhase2.isError():
                    regsCurrentPhase2 = resultCurrentPhase2.registers
#                   print(f"Surova data IL2: {regsCurrentPhase2}")
                    bytes_ = struct.pack('>HH', regsCurrentPhase2[0], regsCurrentPhase2[1])
                    IL2 = struct.unpack('>f', bytes_)[0]
#                   print(f"Float hodnota IL2: {IL2}")
                else:
                    print("Chyba pri cteni:", resultCurrentPhase2)
                client.close()
#               if UL1 > 255:
#                   errorUL1 = "OVERVOLTAGE L1 > 255V"
#               if UL1 < 220:
#                   errorUL1 = "UNDERVOLTAGE L1 < 220V"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultCurrentPhase2 = 0
            IL2 = 302
            errorIL2 = "Invalid data IL2"
            #print("resultCurrentPhase2:", resultCurrentPhase2)
            print("resultCurrentPhase1:", IL2)
#########################################
        try: # CurrentPhase3
            resultCurrentPhase3 = client.read_input_registers(address=10, count=2, slave=1)
            if resultCurrentPhase3.isError():
                resultCurrentPhase3 = 303
#               errorIL3 = "NEZOBRAZUJE SE"
            else:
                resultCurrentPhase3 = client.read_input_registers(address=10, count=2, slave=1)
                if not resultCurrentPhase3.isError():
                    regsCurrentPhase3 = resultCurrentPhase3.registers
#                   print(f"Surova data IL3: {regsCurrentPhase3}")
                    bytes_ = struct.pack('>HH', regsCurrentPhase3[0], regsCurrentPhase3[1])
                    IL3 = struct.unpack('>f', bytes_)[0]
#                   print(f"Float hodnota IL3: {IL3}")
                else:
                    print("Chyba pri cteni:", resultCurrentPhase3)
                client.close()
#               if UL1 > 255:
#                   errorUL1 = "OVERVOLTAGE L1 > 255V"
#               if UL1 < 220:
#                   errorUL1 = "UNDERVOLTAGE L1 < 220V"
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultCurrentPhase3 = 0
            IL3 = 303
            errorIL3 = "Invalid data IL3"
            #print("resultCurrentPhase2:", resultCurrentPhase2)
            print("resultCurrentPhase3:", IL3)
#########################################
        try: # PowerPhase1
            resultPowerPhase1 = client.read_input_registers(address=12, count=2, slave=1)
            if resultPowerPhase1.isError():
                resultPowerPhase1 = 401
                errorPL1 = "NEZOBRAZUJE SE"
            else:
                resultPowerPhase1 = client.read_input_registers(address=12, count=2, slave=1)
                if not resultPowerPhase1.isError():
                    regsPowerPhase1 = resultPowerPhase1.registers
#                   print(f"Surova data: {regsPowerPhase1}")
                    bytes_ = struct.pack('>HH', regsPowerPhase1[0], regsPowerPhase1[1])
                    PL1 = struct.unpack('>f', bytes_)[0]
                    print(f"Float hodnota: {PL1}")
                else:
                    print("Chyba pri cteni:", resultPowerPhase1)
                client.close()
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultPowerPhase1 = 0
            PL1 = 401
            errorPL1 = "Invalid data PL1"
            #print("resultPowerPhase1:", resultPowerPhase1)
            print("resultPowerPhase1:", PL1)
#########################################
        try: # PowerPhase2
            resultPowerPhase2 = client.read_input_registers(address=14, count=2, slave=1)
            if resultPowerPhase2.isError():
                resultPowerPhase2 = 402
                errorPL2 = "NEZOBRAZUJE SE"
            else:
                resultPowerPhase2 = client.read_input_registers(address=14, count=2, slave=1)
                if not resultPowerPhase2.isError():
                    regsPowerPhase2 = resultPowerPhase2.registers
#           		print(f"Surova data: {regsPowerPhase2}")
                    bytes_ = struct.pack('>HH', regsPowerPhase2[0], regsPowerPhase2[1])
                    PL2 = struct.unpack('>f', bytes_)[0]
                    print(f"Float hodnota: {PL2}")
                else:
                    print("Chyba pri cteni:", resultPowerPhase2)
                client.close()
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultPowerPhase2 = 0
            PL2 = 402
            errorPL2 = "Invalid data PL2"
            #print("resultPowerPhase2:", resultPowerPhase2)
            print("resultPowerPhase2:", PL2)
#########################################
        try: # PowerPhase3
            resultPowerPhase3 = client.read_input_registers(address=16, count=2, slave=1)
            if resultPowerPhase3.isError():
                resultPowerPhase3 = 403
                errorPL3 = "NEZOBRAZUJE SE"
            else:
                resultPowerPhase3 = client.read_input_registers(address=16, count=2, slave=1)
                if not resultPowerPhase3.isError():
                    regsPowerPhase3 = resultPowerPhase3.registers
#           		print(f"Surova data: {regsPowerPhase3}")
                    bytes_ = struct.pack('>HH', regsPowerPhase3[0], regsPowerPhase3[1])
                    PL3 = struct.unpack('>f', bytes_)[0]
                    print(f"Float hodnota: {PL3}")
                else:
                    print("Chyba pri cteni:", resultPowerPhase3)
                client.close()
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultPowerPhase3 = 0
            PL3 = 403
            errorPL3 = "Invalid data PL3"
            #print("resultPowerPhase3:", resultPowerPhase3)
            print("resultPowerPhase3:", PL3)
#########################################
        try: # ImportPower
            resultImportPower = client.read_input_registers(address=72, count=2, slave=1)
            if resultImportPower.isError():
                resultImportPower = 500
#               errorPL3 = "NEZOBRAZUJE SE"
            else:
                resultImportPower = client.read_input_registers(address=72, count=2, slave=1)
                if not resultImportPower.isError():
                    regsImportPower = resultImportPower.registers
#           		print(f"Surova data: {regsImportPower}")
                    bytes_ = struct.pack('>HH', regsImportPower[0], regsImportPower[1])
                    impPower = struct.unpack('>f', bytes_)[0]
                else:
                    print("Chyba pri cteni:", resultImportPower)
                client.close()
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultImportPower = 0
            impPower = 500
#            errorPL3 = "Invalid data importPower"
            #print("resultPower:", resultPower)
            print("resultImportPower:", impPower)       
#########################################
        try: # ExportPower
            resultExportPower = client.read_input_registers(address=74, count=2, slave=1)
            if resultExportPower.isError():
                resultExportPower = 600
#               errorPL3 = "NEZOBRAZUJE SE"
            else:
                resultExportPower = client.read_input_registers(address=74, count=2, slave=1)
                if not resultExportPower.isError():
                    regsExportPower = resultExportPower.registers
#           		print(f"Surova data: {regsExportPower}")
                    bytes_ = struct.pack('>HH', regsExportPower[0], regsExportPower[1])
                    expPower = struct.unpack('>f', bytes_)[0]
                else:
                    print("Chyba pri cteni:", resultExportPower)
                client.close()
        except Exception as e:
            print("Chyba pri cteni dat:", e)
            #resultExportPower = 0
            impPower = 600
#           errorPL3 = "Invalid data exportPower"
            #print("resultPower:", resultPower)
            print("resultExportPower:", expPower)            
#########################################
    else:
        print("Nepodarilo se pripojit")
        
######################################### vytvori chybovou zpravu   
    error = (f"{errorUL1} {errorFrequency}")
    
    if (errorFrequency != "" and errorUL1 != "" and errorUL2 != "" and errorUL3 != "" and errorPL1 != ""):
        error = "Energy meter is probably out of the grid!"
        if sendErrorMessageEmoog == 1: # 1 = odesle mail jednou
            subprocess.run(["python3", "emoog.py"]) # odesle mailem chybu energy meter out of grid
            sendErrorMessageEmoog = 0 # 0 = zajisti aby se zprava odeslala jen jednou pri jednom vypadku
            sendGridRecovery = 1
            error = "Grid outage"
######################################### podprogram snizeni vykonu FVE dle casu HDO
# 9:00 - 10:00 casy kdy neni nizky tarif a nenahriha bojler
# 13:00 - 14:00
# 16:00 - 17:00
#    from datetime import datetime, time
#    now = datetime.now().time() # cas HDO
#    
#    if now >= time(9, 0) and now < time(10, 0): # od kdy do kdy bezi nizky tarif
#        if usbrelay0_1ncon == 0:
#            subprocess.run(["sudo", "python3", "usbrelay0_1ncon.py"])
#            usbrelay0_1ncon = 1 # zapamatuje si stav rele, aby se o dva radky vyse na rele neposil povel viz. o jeden radek vyse
#    if now >= time(10, 0) and now < time(10, 0, 30): # v tomto obdobi dojde k prepnuti rele zpet
#        subprocess.run(["sudo", "python3", "usbrelay0_1noon.py"])
#        usbrelay0_1ncon = 0 # a zapise se informace o stavu rele pro dalsi casove pasmo nize
#
#    if now >= time(13, 0) and now < time(14, 0):
#        if usbrelay0_1ncon == 0:
#            subprocess.run(["sudo", "python3", "usbrelay0_1ncon.py"])
#            usbrelay0_1ncon = 1
#    if now >= time(13, 0) and now < time(13, 0, 30):
#        subprocess.run(["sudo", "python3", "usbrelay0_1noon.py"])
#        usbrelay0_1ncon = 0
#    
#    if now >= time(16, 0) and now < time(17, 0):
#        if usbrelay0_1ncon == 0:
#            subprocess.run(["sudo", "python3", "usbrelay0_1ncon.py"])
#            usbrelay0_1ncon = 1
#    if now >= time(17, 0) and now < time(17, 0, 30):
#        subprocess.run(["sudo", "python3", "usbrelay0_1noon.py"])
#        usbrelay0_1ncon = 0
#        
#    if now >= time(23, 25) and now < time(23, 26): # zkusebni podminka
#        if usbrelay0_1ncon == 0:
#            subprocess.run(["sudo", "python3", "usbrelay0_1ncon.py"])
#            usbrelay0_1ncon = 1
#    if now >= time(23, 26) and now < time(23, 26, 30):
#        subprocess.run(["sudo", "python3", "usbrelay0_1noon.py"])
#        usbrelay0_1ncon = 0
######################################### podprogram vycteni cen OTE a vypnuti pretoku pri stanovene cene
    from openpyxl import load_workbook
    import subprocess
    from datetime import datetime

    todayDate = datetime.today()
    todayTime = datetime.today()

    hour = datetime.now().hour
    minute = datetime.now().minute
    day = todayDate.strftime("%d")
    month = todayDate.strftime("%m")
    year = todayDate.strftime("%y")

#   if minute == 1 or minute == 15 or minute == 30 or minute == 45: # ve zvolene minute spusti proces nacteni cen pro aktualni den
    if minute ==  1 and hour == 0: # ve stanoveny cas nacte data o cenach OTE, aby to zbytecne nenacital kazdy cyklus
        wb = load_workbook(f"/home/tomas/OTE/OTEfiles/DT_{day}_{month}_20{year}_CZ.xlsx")
        sheet = wb.active

        cellsPrice = ['B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B30', 'B31', 'B32', 'B33', 'B34', 'B35',
                    'B36', 'B37', 'B38', 'B39', 'B40', 'B41', 'B42', 'B43', 'B44', 'B45', 'B46', 'B47']

        #cellsHour = ['A24', 'A25', 'A26', 'A27', 'A28', 'A29', 'A30', 'A31', 'A32', 'A33', 'A34', 'A35',
#            	'A36', 'A37', 'A38', 'A39', 'A40', 'A41', 'A42', 'A43', 'A45', 'A47']

        seznamPrice = [sheet[cell].value for cell in cellsPrice]
        #seznamHour = [sheet[cell].value for cell in cellsHour]

        #print(seznamPrice)
        #print(seznamHour)

#    	if minute == 12: # ve zvolene minute spusti proces nacteni cen pro aktualni den
    for i in range(0, 24): # nacte seznam cen pro kazdou hodinu aktualniho dne
        a = i + 1 # posun o 1 aby sedeli data v bunce A24 vs. B24
        #print(a)
        #print(seznamPrice[i]) # tiskne cenu pro konkretni hodinu
            
        if hour == a: # ulozi cenu OTE pro DB
            priceOTE = seznamPrice[i]
    
        if statusRelay == False and hour == a and seznamPrice[i] < 10: # nastaveni ceny pri ktere se ma omezit pretok
            print(f"Cena v hodinu {a} je nizsi nez 10EUR/MWh")
#           priceOTE = seznamPrice[i]
            print(f"Cena OTE pro aktualni hodinu je {priceOTE}EUR/MWh")
            #subprocess.run(["sudo", "python3", "/home/tomas/OTE/usbrelay0_1ncon.py"])
            count = usbrelay_py.board_count() # ver.15 zmena volani funkce spinani rele
            print("Count: ",count)
            boards = usbrelay_py.board_details()
            for board in boards:
                result = usbrelay_py.board_control(board[0],relay,1)
            statusRelay = True
            print(statusRelay)
        
        if statusRelay == True and hour == a and seznamPrice[i] >= 10: # rozepnuti rele
            #subprocess.run(["sudo", "python3", "/home/tomas/OTE/usbrelay0_1noon.py"])
            count = usbrelay_py.board_count() # ver.15 zmena volani funkce spinani rele
            print("Count: ",count)
            boards = usbrelay_py.board_details()
            for board in boards:
                result = usbrelay_py.board_control(board[0],relay,0)
            statusRelay = False
            print(statusRelay)
    
    print(statusRelay)
    print(priceOTE)
    
#   if minute == 59 or minute == 14 or minute == 29 or minute == 44: # ve zvolene minute vypne rele ##### prostor pro vyladeni, aby se celou minutu kazdou hodinu neposilal povel k vypnuti, ale jen jednou
#       subprocess.run(["sudo", "python3", "usbrelay0_1noon.py"])
#       statusRelay60perc = False
#       print(statusRelay60perc)
                
######################################### zapis do databaze
    conn = sqlite3.connect("/home/tomas/OTE/test11.db")
    c = conn.cursor()
    date = datetime.now().strftime("%Y-%m-%d")
    time = datetime.now().strftime("%H:%M:%S")
    c.execute("INSERT INTO cisla (date, time, frequency, UL1, UL2, UL3, IL1, IL2, IL3, PL1, PL2, PL3, impPower, expPower, priceOTE, statusRelay, error) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (date, time, frequency, UL1, UL2, UL3, IL1, IL2, IL3, PL1, PL2, PL3, impPower, expPower, priceOTE, statusRelay, error))
    conn.commit()
    conn.close()
######################################### vynuluje predchozi error
    
    error = ""
    errorFrequency = ""
    errorUL1 = ""
    errorUL2 = ""
    errorUL3 = ""
    errorIL1 = ""
    errorIL2 = ""
    errorIL3 = ""
    errorPL1 = ""
    errorPL2 = ""
    errorPL3 = ""
    impPower = ""
    expPower = ""
    
######################################### nastaveni timeframe vycitani dat
    
    import time
    time.sleep(5)

######################################### nacteni databaze
#   conn = sqlite3.connect("test33.db")
#   c = conn.cursor()
#   c.execute("SELECT * FROM cisla")
#   print(c.fetchall())
#   conn.close()
#########################################
    